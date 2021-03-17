import openpyxl;
import os.path;
import time;

class GetInputsClass:
	def __init__(self, pathToCheck):
		self.pathToCheck = pathToCheck;
		self.pathHasFile = False;
		self.excelWorkBook = None;
		self.excelWorkSheet = None;
		self.keywords = [];

	async def check_if_path_has_file(self, pathToCheck):
		if(os.path.isfile(pathToCheck)):
			return True;
		return False;

	async def get_the_work_book(self, pathToCheck):
		excelWorkBook = openpyxl.load_workbook(pathToCheck);
		return excelWorkBook;

	async def get_the_active_work_sheet(self, excelWorkBook):
		excelWorkSheet = excelWorkBook.active;
		return excelWorkSheet;

	async def get_inputs_from_work_sheet(self, excelWorkSheet):
		keywords = [];
		rowsCount = excelWorkSheet.max_row;
		colsCount = excelWorkSheet.max_column;
		print('colsCount', colsCount);
		print('rowsCount', rowsCount);
		for row in range(2,(rowsCount+1),1):
			keywords.append(excelWorkSheet.cell(row, 1).value);
		return keywords;

	async def get_keywords(self):
		return self.keywords;

	async def execute(self):
		print("path to look at for input file is -- {0}".format(self.pathToCheck));
		print("checking if the file exists or not at given location");

		self.pathHasFile = await self.check_if_path_has_file(self.pathToCheck);
		print('pathHasFile', self.pathHasFile);

		if (self.pathHasFile):
			self.excelWorkBook = await self.get_the_work_book(self.pathToCheck);

		if(self.excelWorkBook != None):
			print('got the workbook -- now need to get the sheet which contains inputs');
			print('getting the first active sheet');
			self.excelWorkSheet = await self.get_the_active_work_sheet(self.excelWorkBook);

		if(self.excelWorkSheet != None):
			print('got the worksheet!!', type(self.excelWorkSheet));
			print('need to extract the keywords');
			self.keywords = await self.get_inputs_from_work_sheet(self.excelWorkSheet);

		if(len(self.keywords) > 0):
			print('got a total of {0} keywords from the file given'.format(len(self.keywords)));
			return True;
		else:
			print('got these many keywords - {0}'.format(len(self.keywords)));

		return False;

